import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

"""
Recommended LibreOffice
"""


# TODO parametrize & create a function
# TODO read sheet, col ids and cell styles (centered, left, colored, character if true, etc) from yaml

MASVS_TITLES = {
    'V1': 'Architecture, Design and Threat Modeling Requirements',
    'V2': 'Data Storage and Privacy Requirements',
    'V3': 'Cryptography Requirements',
    'V4': 'Authentication and Session Management Requirements',
    'V5': 'Network Communication Requirements',
    'V6': 'Platform Interaction Requirements',
    'V7': 'Code Quality and Build Setting Requirements',
    'V8': 'Resilience Requirements',
}

def get_hyperlink(url):
    title = url.split('#')[1].replace('-',' ').capitalize() 
    return f'=HYPERLINK("{url}", "{title}")'

CHECKMARK = "✓"

def write_table(masvs_file, input_file, output_file):

    masvs_dict = yaml.safe_load(open(masvs_file))

    wb = load_workbook(filename=input_file)

    from openpyxl.styles import NamedStyle, Font, Border, Side
    centered = NamedStyle(name="centered")
    centered.font = Font(name='Calibri')
    bd = Side(style='thick', color="FFFFFF")
    centered.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    centered.alignment = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)

    wb.add_named_style(centered)

    # table_config = {
    #     'sheet': 'Security Requirements - Android'
    #     'start_row': 4
    #     'start_col': 2
    #     'columns': [
    #         {'name': 'ID',},
    #         {'name': 'MSTG-ID',},
    #         {'name': 'Detailed Verification Requirement',},
    #         {'name': 'L1', 'style': 'centered', 'special_char': "✓"},
    #         {'name': 'L2', 'style': 'centered',  'special_char': "✓"},
    #         {'name': 'R', 'style': 'centered',  'special_char': "✓"},
    #         {'name': 'References',},
    #     ]
            
    # }

    table = wb['Security Requirements - Android']

    row=4
    col_id=2
    col_mstg_id=3
    col_text=4
    col_l1=5
    col_l2=6
    col_r=7
    col_link=8

    for mstg_id, req in masvs_dict.items():
        req_id = req['id'].split('.') 
        category = req_id[0]
        subindex = req_id[1]

        if subindex == '1':
            category_id = f"V{category}"
            category_title = MASVS_TITLES[category_id]
            table.cell(row=row,column=col_id).value = category_id
            table.cell(row=row,column=col_id+2).value = category_title
            row = row+1
        
        l1 = CHECKMARK if req['L1'] else ""
        l2 = CHECKMARK if req['L2'] else ""
        r = CHECKMARK if req['R'] else ""

        table.cell(row=row,column=col_id).value = req['id']
        table.cell(row=row,column=col_mstg_id).value = mstg_id
        table.cell(row=row,column=col_text).value = req['text']
        table.cell(row=row,column=col_text).style = 'centered' 
        
        if l1 != "":
            table.cell(row=row,column=col_l1).value = l1
            table.cell(row=row,column=col_l1).fill = PatternFill("solid", fgColor="0033CCCC")
        if l2 != "":
            table.cell(row=row,column=col_l2).value = l2
            table.cell(row=row,column=col_l2).fill = PatternFill("solid", fgColor="0099CC00")
        if r != "":
            table.cell(row=row,column=col_r).value = r
            table.cell(row=row,column=col_r).fill = PatternFill("solid", fgColor="00FF9900")
        if req.get('links'):
            table.cell(row=row,column=col_link).value = get_hyperlink(req['links'][0])
        
        table.row_dimensions[row].height = 45 # points

        row = row+1

    align_center = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)
    align_left = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0, justifyLastLine=True)

    table.sheet_view.showGridLines = False

    for row in table.iter_rows():
        for cell in row:
            cell.alignment = align_center

    for cell in table['D']:
        cell.alignment = align_left

    wb.save(filename=output_file)

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Export the MASVS requirements as Excel. Default language is en.')
    parser.add_argument('-m', '--masvs', required=True)
    parser.add_argument('-i', '--inputfile', required=True)
    parser.add_argument('-o', '--outputfile', required=True)

    args = parser.parse_args()

    write_table(args.masvs, args.inputfile, args.outputfile)


if __name__ == '__main__':
    main()